#!/usr/bin/python
# -*- coding: utf-8 -*-
from PyQt5 import QtWidgets, uic
import pyodbc
import openpyxl
from openpyxl import Workbook
import sys
from lkm_gui import Ui_MainWindow
import sql_querys
from PyQt5.QtWidgets import QMainWindow, QApplication


class SpareParts(QtWidgets.QMainWindow):
    wb = Workbook()
    server = 'host'
    database = 'db'
    username = 'sa'
    password = 'pw'
    driver = '{SQL Server}'  # Driver you need to connect to the database
    port = '1433'

    def __init__(self):
        self.ui = Ui_MainWindow()
        super().__init__()
        self.ui.setupUi(self)
        self.ui.selectButton.clicked.connect(lambda x: self.showDialog())
        self.ui.save.clicked.connect(lambda x: self.start_main_work())
        self.cnn = pyodbc.connect(
            'DRIVER=' + self.driver + ';PORT=port;SERVER=' + self.server + ';PORT=1443;DATABASE=' + self.database + ';UID=' + self.username +
            ';PWD=' + self.password)
        self.count_list = list()
        self.code_list = list()
        self.provider_list = list()
        self.STOCKID_list = list()
        self.cursor = self.cnn.cursor()


    def showDialog(self):
        # self.clear_all_lists()
        fname = QtWidgets.QFileDialog.getOpenFileName(self, 'Open file', '*.xlsm')[0]
        name_index_ = fname.rfind("/")
        self.filename = fname
        self.ui.incoming_file_name.setText(fname[name_index_ + 1:])

    def create_lists(self, later, row_num, lst):
        row_max = self.ws.max_row
        while int(row_num) <= row_max and self.ws[f"{later}{row_num}"].value is not None:
            lst.append(self.ws[f"{later}{row_num}"].value)
            row_num = int(row_num) + 1
        try:
            self.insert_into_inventory_latter()
            self.ui.res.setText("Импорт успешен")
            self.ui.res.setStyleSheet("color : green")

        except Exception as err:
            self.ui.res.setText(f"Обнаружены ошибки, {err}")
            self.ui.res.setStyleSheet("color : red")

    def insert_into_inventory_latter(self):
        self.ui.progressBar.setMaximum(len(self.code_list))
        self.cnt = 0
        for code, count, suplno, stockid in zip(self.code_list, self.count_list, self.provider_list, self.STOCKID_list):
            self.cursor.execute(sql_querys.insert_into_invt1(inventory_list_number=self.ui.inventory_list_number.text(),
                                                             code=code, count=count, SUPLNO=suplno, STOCKID=stockid))
            print(sql_querys.insert_into_invt1(inventory_list_number=self.ui.inventory_list_number.text(),
                                         code=code, count=count, SUPLNO=suplno, STOCKID=stockid))
            self.cnn.commit()

            self.cnt += 1
            self.ui.progressBar.setValue(self.cnt)


    def start_main_work(self):
        print("start")
        self.wb = openpyxl.load_workbook(self.filename, data_only=True)
        self.ws = self.wb[self.ui.lineEdit_4.text()]
        self.create_lists(self.ui.lineEdit.text()[:1], self.ui.lineEdit.text()[1:], self.code_list)
        self.create_lists(self.ui.lineEdit_3.text()[:1], self.ui.lineEdit_3.text()[1:], self.count_list)
        self.create_lists(self.ui.provider.text()[:1], self.ui.provider.text()[1:], self.provider_list)
        self.create_lists(self.ui.stockid.text()[:1], self.ui.stockid.text()[1:], self.STOCKID_list)




def main():
    app = QApplication(sys.argv)
    w = SpareParts()
    w.show()
    app.exec_()


if __name__ == '__main__':
    main()
